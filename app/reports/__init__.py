import io
from datetime import date, datetime
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file, jsonify
from flask_login import login_required, current_user
from functools import wraps
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from sqlalchemy import func, and_
from app import db
from app.models import (
    Request, Material, InventorySnapshot, IssuedMaterial,
    Approval, User, ReportLog, AuditLog
)

reports = Blueprint('reports', __name__)


def store_or_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not (
            current_user.is_store_manager() or current_user.is_admin()
        ):
            flash('Access denied.', 'danger')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function


@reports.route('/dashboard')
@login_required
@store_or_admin_required
def dashboard():
    report_date_str = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
    except ValueError:
        report_date = date.today()

    stats = get_report_stats(report_date)
    recent_reports = ReportLog.query.order_by(ReportLog.created_at.desc()).limit(8).all()

    # Issued items for display
    issued_requests = Request.query.filter(
        func.date(Request.created_at) == report_date,
        Request.status == 'Issued'
    ).order_by(Request.created_at.desc()).all()

    # All requests for the requests table (last 10)
    all_requests = Request.query.filter(
        func.date(Request.created_at) == report_date
    ).order_by(Request.created_at.desc()).limit(20).all()

    return render_template('reports/dashboard.html',
                           stats=stats,
                           report_date=report_date,
                           recent_reports=recent_reports,
                           issued_requests=issued_requests,
                           all_requests=all_requests)


def get_report_stats(report_date=None):
    query = Request.query
    if report_date:
        query = query.filter(func.date(Request.created_at) == report_date)

    all_requests = query.all()
    total = len(all_requests)
    approved = sum(1 for r in all_requests if r.status == 'Approved')
    rejected = sum(1 for r in all_requests if r.status == 'Rejected')
    pending = sum(1 for r in all_requests if r.status == 'Pending Approval')
    issued = sum(1 for r in all_requests if r.status == 'Issued')
    total_value = sum(float(r.amount) for r in all_requests)
    issued_value = sum(float(r.amount) for r in all_requests if r.status == 'Issued')

    return {
        'total': total,
        'approved': approved,
        'rejected': rejected,
        'pending': pending,
        'issued': issued,
        'total_value': total_value,
        'issued_value': issued_value,
    }


@reports.route('/download')
@login_required
@store_or_admin_required
def download_report():
    report_date_str = request.args.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
    except ValueError:
        report_date = date.today()

    wb = generate_excel_report(report_date)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Engineering_Stores_Report_{report_date.strftime('%Y%m%d')}.xlsx"

    # Log report generation
    log = ReportLog(
        report_type='Daily Report',
        report_date=report_date,
        generated_by=current_user.id,
        filename=filename
    )
    db.session.add(log)
    db.session.commit()

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def style_header(ws, row_num, col_count):
    """Apply header styling to a row."""
    header_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def auto_size_columns(ws):
    """Auto-size columns based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 4, 60)
        ws.column_dimensions[col_letter].width = adjusted_width


def generate_excel_report(report_date):
    wb = Workbook()

    # Colors
    title_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
    title_font = Font(color='FFFFFF', bold=True, size=14)
    sub_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
    alt_fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Fetch all data for the date
    all_requests = Request.query.filter(
        func.date(Request.created_at) == report_date
    ).all()

    # ──────────────────────────────────────────────
    # SHEET 1: Executive Summary
    # ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Executive Summary'

    ws1.merge_cells('A1:D1')
    c = ws1['A1']
    c.value = 'ENGINEERING STORES MANAGEMENT SYSTEM'
    c.font = Font(bold=True, size=16, color='1B3A5C')
    c.alignment = Alignment(horizontal='center')
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells('A2:D2')
    c = ws1['A2']
    c.value = f'Daily Report - {report_date.strftime("%d %B %Y")}'
    c.font = Font(bold=True, size=13, color='4A4A4A')
    c.alignment = Alignment(horizontal='center')
    ws1.row_dimensions[2].height = 25

    ws1.merge_cells('A3:D3')
    c = ws1['A3']
    c.value = f'Generated on: {datetime.utcnow().strftime("%d-%b-%Y %H:%M")} UTC'
    c.font = Font(italic=True, size=10, color='888888')
    c.alignment = Alignment(horizontal='center')
    ws1.row_dimensions[3].height = 20

    # Metrics
    stats = get_report_stats(report_date)
    metrics = [
        ['Metric', 'Value'],
        ['Total Requests', stats['total']],
        ['Approved Requests', stats['approved']],
        ['Rejected Requests', stats['rejected']],
        ['Pending Requests', stats['pending']],
        ['Issued Requests', stats['issued']],
        ['Total Value Requested', f"Rs. {stats['total_value']:,.2f}"],
        ['Total Value Issued', f"Rs. {stats['issued_value']:,.2f}"],
    ]

    for i, row_data in enumerate(metrics):
        row_num = i + 5
        for col_num, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if i == 0:
                cell.fill = title_fill
                cell.font = Font(color='FFFFFF', bold=True, size=11)
                cell.alignment = Alignment(horizontal='center')
            elif i % 2 == 0:
                cell.fill = alt_fill
            if col_num == 1:
                cell.font = Font(bold=True)

    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 25

    # ──────────────────────────────────────────────
    # SHEET 2: All Requests
    # ──────────────────────────────────────────────
    ws2 = wb.create_sheet('All Requests')
    headers = ['Request No.', 'Requester', 'Department', 'Material Code', 'Material Description',
               'UOM', 'Quantity', 'Unit Rate', 'Amount', 'Cost Center', 'GL Code', 'Status', 'Created At']
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, req in enumerate(all_requests, 2):
        fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid') if i % 2 == 0 else None
        row_data = [
            req.request_number, req.requester_name, req.department,
            req.material_code, req.material_description, req.uom,
            float(req.quantity_required) if req.quantity_required else 0,
            float(req.unit_rate) if req.unit_rate else 0,
            float(req.amount) if req.amount else 0,
            req.cost_center, req.gl_code, req.status,
            req.created_at.strftime('%d-%b-%Y %H:%M') if req.created_at else ''
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if fill:
                cell.fill = fill
    auto_size_columns(ws2)

    # ──────────────────────────────────────────────
    # SHEETS 3-5: Filtered by status
    # ──────────────────────────────────────────────
    for sheet_name, status_filter in [
        ('Approved Requests', 'Approved'),
        ('Rejected Requests', 'Rejected'),
        ('Pending Requests', 'Pending Approval')
    ]:
        ws = wb.create_sheet(sheet_name)
        filtered = [r for r in all_requests if r.status == status_filter]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for i, req in enumerate(filtered, 2):
            fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid') if i % 2 == 0 else None
            row_data = [
                req.request_number, req.requester_name, req.department,
                req.material_code, req.material_description, req.uom,
                float(req.quantity_required) if req.quantity_required else 0,
                float(req.unit_rate) if req.unit_rate else 0,
                float(req.amount) if req.amount else 0,
                req.cost_center, req.gl_code, req.status,
                req.created_at.strftime('%d-%b-%Y %H:%M') if req.created_at else ''
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin_border
                if fill:
                    cell.fill = fill
        auto_size_columns(ws)

    # ──────────────────────────────────────────────
    # SHEET 6: Issued Materials
    # ──────────────────────────────────────────────
    ws6 = wb.create_sheet('Issued Materials')
    issued_headers = ['Request No.', 'Requester', 'Department', 'Material Code', 'Material Description',
                      'UOM', 'Quantity', 'Amount', 'Issued By', 'Issue Date', 'Issue Time']
    for col, h in enumerate(issued_headers, 1):
        cell = ws6.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    issued = [r for r in all_requests if r.status == 'Issued']
    for i, req in enumerate(issued, 2):
        fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid') if i % 2 == 0 else None
        issuer = User.query.get(req.issued_material.issued_by) if req.issued_material else None
        row_data = [
            req.request_number, req.requester_name, req.department,
            req.material_code, req.material_description, req.uom,
            float(req.quantity_required) if req.quantity_required else 0,
            float(req.amount) if req.amount else 0,
            issuer.full_name if issuer else '',
            str(req.issued_material.issued_date) if req.issued_material else '',
            str(req.issued_material.issued_time) if req.issued_material else '',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws6.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if fill:
                cell.fill = fill
    auto_size_columns(ws6)

    # ──────────────────────────────────────────────
    # SHEET 7: Material Consumption Summary
    # ──────────────────────────────────────────────
    ws7 = wb.create_sheet('Material Consumption')
    mat_headers = ['Material Code', 'Material Description', 'UOM', 'Qty Requested', 'Qty Issued', 'Amount']
    for col, h in enumerate(mat_headers, 1):
        cell = ws7.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    mat_summary = {}
    for req in all_requests:
        key = req.material_code
        if key not in mat_summary:
            mat_summary[key] = {
                'desc': req.material_description,
                'uom': req.uom,
                'qty_req': 0,
                'qty_issued': 0,
                'amount': 0
            }
        mat_summary[key]['qty_req'] += float(req.quantity_required) if req.quantity_required else 0
        mat_summary[key]['amount'] += float(req.amount) if req.amount else 0
        if req.status == 'Issued':
            mat_summary[key]['qty_issued'] += float(req.quantity_required) if req.quantity_required else 0

    for i, (code, data) in enumerate(mat_summary.items(), 2):
        fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid') if i % 2 == 0 else None
        row_data = [code, data['desc'], data['uom'], data['qty_req'], data['qty_issued'], data['amount']]
        for col, val in enumerate(row_data, 1):
            cell = ws7.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if fill:
                cell.fill = fill
    auto_size_columns(ws7)

    # ──────────────────────────────────────────────
    # SHEET 8: Department Consumption
    # ──────────────────────────────────────────────
    ws8 = wb.create_sheet('Dept Consumption')
    dept_headers = ['Department', 'Total Requests', 'Total Amount']
    for col, h in enumerate(dept_headers, 1):
        cell = ws8.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    dept_summary = {}
    for req in all_requests:
        dept = req.department or 'Unknown'
        if dept not in dept_summary:
            dept_summary[dept] = {'count': 0, 'amount': 0}
        dept_summary[dept]['count'] += 1
        dept_summary[dept]['amount'] += float(req.amount) if req.amount else 0

    for i, (dept, data) in enumerate(dept_summary.items(), 2):
        fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid') if i % 2 == 0 else None
        row_data = [dept, data['count'], data['amount']]
        for col, val in enumerate(row_data, 1):
            cell = ws8.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if fill:
                cell.fill = fill
    auto_size_columns(ws8)

    # ──────────────────────────────────────────────
    # SHEET 9: Cost Center Summary
    # ──────────────────────────────────────────────
    ws9 = wb.create_sheet('Cost Center Summary')
    cc_headers = ['Cost Center', 'Total Requests', 'Total Amount']
    for col, h in enumerate(cc_headers, 1):
        cell = ws9.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    cc_summary = {}
    for req in all_requests:
        cc = req.cost_center or 'Not Specified'
        if cc not in cc_summary:
            cc_summary[cc] = {'count': 0, 'amount': 0}
        cc_summary[cc]['count'] += 1
        cc_summary[cc]['amount'] += float(req.amount) if req.amount else 0

    for i, (cc, data) in enumerate(cc_summary.items(), 2):
        fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid') if i % 2 == 0 else None
        row_data = [cc, data['count'], data['amount']]
        for col, val in enumerate(row_data, 1):
            cell = ws9.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if fill:
                cell.fill = fill
    auto_size_columns(ws9)

    # ──────────────────────────────────────────────
    # SHEET 10: GL Code Summary
    # ──────────────────────────────────────────────
    ws10 = wb.create_sheet('GL Code Summary')
    gl_headers = ['GL Code', 'Description', 'Total Amount']
    for col, h in enumerate(gl_headers, 1):
        cell = ws10.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    gl_summary = {}
    for req in all_requests:
        gl = req.gl_code or 'Not Specified'
        if gl not in gl_summary:
            gl_summary[gl] = {'desc': '', 'amount': 0}
        gl_summary[gl]['amount'] += float(req.amount) if req.amount else 0

    for i, (gl, data) in enumerate(gl_summary.items(), 2):
        fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid') if i % 2 == 0 else None
        row_data = [gl, data['desc'], data['amount']]
        for col, val in enumerate(row_data, 1):
            cell = ws10.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if fill:
                cell.fill = fill
    auto_size_columns(ws10)

    # ──────────────────────────────────────────────
    # SHEET 11: Approval Audit Trail
    # ──────────────────────────────────────────────
    ws11 = wb.create_sheet('Approval Audit Trail')
    audit_headers = ['Request No.', 'Approver', 'Action', 'Remarks', 'Actioned At']
    for col, h in enumerate(audit_headers, 1):
        cell = ws11.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    all_approvals = Approval.query.join(Request).filter(
        func.date(Request.created_at) == report_date
    ).order_by(Approval.actioned_at).all()

    for i, appr in enumerate(all_approvals, 2):
        fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid') if i % 2 == 0 else None
        row_data = [
            appr.request.request_number if appr.request else '',
            appr.approver.full_name if appr.approver else '',
            appr.action,
            appr.remarks or '',
            appr.actioned_at.strftime('%d-%b-%Y %H:%M') if appr.actioned_at else ''
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws11.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if fill:
                cell.fill = fill
    auto_size_columns(ws11)

    # ──────────────────────────────────────────────
    # SHEET 12: Inventory Snapshot
    # ──────────────────────────────────────────────
    ws12 = wb.create_sheet('Inventory Snapshot')
    inv_headers = ['Material Code', 'Material Description', 'UOM', 'Unit Rate', 'Available Stock', 'GL Code', 'Cost Center']
    for col, h in enumerate(inv_headers, 1):
        cell = ws12.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    active_snapshot = InventorySnapshot.query.filter_by(is_active=True).first()
    if active_snapshot:
        materials = Material.query.filter_by(snapshot_id=active_snapshot.id).all()
        for i, mat in enumerate(materials, 2):
            fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid') if i % 2 == 0 else None
            row_data = [
                mat.material_code, mat.material_description, mat.uom,
                float(mat.unit_rate) if mat.unit_rate else 0,
                float(mat.available_stock) if mat.available_stock else 0,
                mat.gl_code, mat.cost_center
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws12.cell(row=i, column=col, value=val)
                cell.border = thin_border
                if fill:
                    cell.fill = fill
    auto_size_columns(ws12)

    return wb

from flask import Response
import csv
from datetime import timedelta

from sqlalchemy.orm import joinedload

def build_detailed_query(request_args):
    """Helper to build the base query from filters."""
    query = Request.query.options(
        joinedload(Request.requester),
        joinedload(Request.material),
        joinedload(Request.issued_material).joinedload(IssuedMaterial.issued_by_user),
        joinedload(Request.approvals).joinedload(Approval.approver)
    ).join(User, Request.user_id == User.id)
    
    # 1. Date Filter (Default: Last 30 days)
    start_date_str = request_args.get('start_date')
    end_date_str = request_args.get('end_date')
    
    if not start_date_str and not end_date_str:
        end_date = date.today()
        start_date = end_date - timedelta(days=30)
    else:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else date(2000, 1, 1)
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else date.today()
        except ValueError:
            end_date = date.today()
            start_date = end_date - timedelta(days=30)
            
    query = query.filter(func.date(Request.created_at) >= start_date)
    query = query.filter(func.date(Request.created_at) <= end_date)
    
    # 2. Status Filter
    status = request_args.get('status')
    if status and status != 'All':
        query = query.filter(Request.status == status)
        
    # 3. Employee Filter
    employee = request_args.get('employee', '').strip()
    if employee:
        query = query.filter(
            db.or_(
                User.full_name.ilike(f'%{employee}%'),
                User.employee_id.ilike(f'%{employee}%'),
                User.username.ilike(f'%{employee}%')
            )
        )
        
    # 4. Material Code Filter
    material_code = request_args.get('material_code', '').strip()
    if material_code:
        query = query.filter(Request.material_code.ilike(f'%{material_code}%'))

    # Sort
    query = query.order_by(Request.created_at.desc())
    
    return query, start_date, end_date

def get_detailed_totals(base_query):
    all_reqs = base_query.all()
    totals = {
        'total_requests': len(all_reqs),
        'total_approved': sum(1 for r in all_reqs if r.status in ('Approved', 'Issued')),
        'total_rejected': sum(1 for r in all_reqs if r.status == 'Rejected'),
        'total_issued': sum(1 for r in all_reqs if r.status == 'Issued'),
        'qty_requested': sum(float(r.quantity_required or 0) for r in all_reqs),
        'qty_issued': sum(float(r.quantity_required or 0) for r in all_reqs if r.status == 'Issued')
    }
    # Wait, the prompt says "Total Approved". Should that include Issued?
    # Usually "Approved" implies currently approved, or ever approved. Let's strictly use statuses.
    # Actually, Issued requests *were* approved. Let's just do status == 'Approved'.
    totals['total_approved'] = sum(1 for r in all_reqs if r.status == 'Approved')
    
    return totals

def format_detailed_row(req):
    approver_name = '-'
    approval_date = '-'
    # Get the latest approval from the preloaded list
    if req.approvals:
        sorted_approvals = sorted(req.approvals, key=lambda a: a.actioned_at, reverse=True)
        latest_approval = sorted_approvals[0]
        if latest_approval.approver:
            approver_name = latest_approval.approver.full_name
            approval_date = latest_approval.actioned_at.strftime('%Y-%m-%d %H:%M') if latest_approval.actioned_at else '-'
        
    issuer_name = '-'
    issue_date = '-'
    if req.issued_material:
        issuer_name = req.issued_material.issued_by_user.full_name if req.issued_material.issued_by_user else '-'
        issue_date = req.issued_material.issued_date.strftime('%Y-%m-%d') if req.issued_material.issued_date else '-'
        
    req_value = (float(req.quantity_required or 0) * float(req.unit_rate or 0))

    return {
        'request_number': req.request_number,
        'request_date': req.created_at.strftime('%Y-%m-%d'),
        'employee_name': req.requester_name,
        'employee_id': req.requester.employee_id if req.requester else '-',
        'material_code': req.material_code,
        'material_description': req.material_description,
        'quantity_requested': float(req.quantity_required or 0),
        'uom': req.uom,
        'request_value': req_value,
        'cost_center': req.cost_center or '-',
        'gl_code': req.gl_code or '-',
        'status': req.status,
        'approver_name': approver_name,
        'approval_date': approval_date,
        'issuer_name': issuer_name,
        'issue_date': issue_date
    }


@reports.route('/detailed')
@login_required
@store_or_admin_required
def detailed_report():
    query, start_date, end_date = build_detailed_query(request.args)
    totals = get_detailed_totals(query)
    
    page = request.args.get('page', 1, type=int)
    pagination = query.paginate(page=page, per_page=50, error_out=False)
    
    # Format rows for template
    formatted_rows = [format_detailed_row(req) for req in pagination.items]
    
    return render_template('reports/detailed.html', 
                           rows=formatted_rows, 
                           pagination=pagination,
                           totals=totals,
                           start_date=start_date.strftime('%Y-%m-%d'),
                           end_date=end_date.strftime('%Y-%m-%d'),
                           status_filter=request.args.get('status', 'All'),
                           employee_filter=request.args.get('employee', ''),
                           material_filter=request.args.get('material_code', ''))

@reports.route('/detailed/export')
@login_required
@store_or_admin_required
def export_detailed_report():
    query, start_date, end_date = build_detailed_query(request.args)
    all_requests = query.all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        'Request Number', 'Request Date', 'Employee Name', 'Employee ID',
        'Material Code', 'Material Description', 'Quantity Requested', 'UOM',
        'Request Value', 'Cost Center', 'GL Code', 'Status', 
        'Approver Name', 'Approval Date', 'Store Manager Name', 'Issue Date'
    ])
    
    for req in all_requests:
        row = format_detailed_row(req)
        writer.writerow([
            row['request_number'], row['request_date'], row['employee_name'], row['employee_id'],
            row['material_code'], row['material_description'], row['quantity_requested'], row['uom'],
            row['request_value'], row['cost_center'], row['gl_code'], row['status'],
            row['approver_name'], row['approval_date'], row['issuer_name'], row['issue_date']
        ])
        
    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers['Content-Disposition'] = f'attachment; filename=detailed_report_{datetime.utcnow().strftime("%Y%m%d%H%M")}.csv'
    return response
@reports.route('/detailed/export/excel')
@login_required
@store_or_admin_required
def export_detailed_report_excel():
    query, start_date, end_date = build_detailed_query(request.args)
    all_requests = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Report"
    
    headers = [
        'Request Number', 'Request Date', 'Employee Name', 'Employee ID',
        'Material Code', 'Material Description', 'Quantity Requested', 'UOM',
        'Request Value', 'Cost Center', 'GL Code', 'Status', 
        'Approver Name', 'Approval Date', 'Store Manager Name', 'Issue Date'
    ]
    
    # Apply header styles
    header_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        
    for i, req in enumerate(all_requests, 2):
        row_dict = format_detailed_row(req)
        fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid') if i % 2 == 0 else None
        
        row_data = [
            row_dict['request_number'], row_dict['request_date'], row_dict['employee_name'], row_dict['employee_id'],
            row_dict['material_code'], row_dict['material_description'], row_dict['quantity_requested'], row_dict['uom'],
            row_dict['request_value'], row_dict['cost_center'], row_dict['gl_code'], row_dict['status'],
            row_dict['approver_name'], row_dict['approval_date'], row_dict['issuer_name'], row_dict['issue_date']
        ]
        
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if fill:
                cell.fill = fill
                
    auto_size_columns(ws)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"detailed_report_{datetime.utcnow().strftime('%Y%m%d%H%M')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
