"""
Script to create a sample inventory Excel file for testing uploads.
Run: python create_sample_inventory.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Inventory"

# Headers
headers = ['Material Code', 'Material Description', 'UOM', 'Unit Rate', 'Available Stock', 'GL Code', 'Cost Center']
header_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Sample data
data = [
    ('MAT-001', 'MS Angle 50x50x5mm', 'KG', 85.50, 500.0, 'GL-5001', 'CC-MECH'),
    ('MAT-002', 'GI Pipe 1 inch', 'MTR', 145.00, 200.0, 'GL-5002', 'CC-ELEC'),
    ('MAT-003', 'Bearing 6205 ZZ', 'NOS', 320.00, 50.0, 'GL-5003', 'CC-MAINT'),
    ('MAT-004', 'Allen Key Set (7pcs)', 'SET', 250.00, 30.0, 'GL-5001', 'CC-MECH'),
    ('MAT-005', 'V-Belt A-45', 'NOS', 180.00, 100.0, 'GL-5003', 'CC-MAINT'),
    ('MAT-006', 'Copper Wire 1.5 sqmm (100m)', 'ROLL', 1250.00, 25.0, 'GL-5002', 'CC-ELEC'),
    ('MAT-007', 'MS Hex Bolt M12x50', 'KG', 75.00, 300.0, 'GL-5001', 'CC-MECH'),
    ('MAT-008', 'Grease NLGI 2 (500g)', 'TIN', 95.00, 80.0, 'GL-5003', 'CC-MAINT'),
    ('MAT-009', 'PVC Conduit 25mm (3m)', 'NOS', 55.00, 150.0, 'GL-5002', 'CC-ELEC'),
    ('MAT-010', 'Cutting Wheel 4 inch', 'NOS', 45.00, 200.0, 'GL-5004', 'CC-WELDING'),
    ('MAT-011', 'MS Sheet 3mm (1220x2440)', 'NOS', 4500.00, 20.0, 'GL-5001', 'CC-MECH'),
    ('MAT-012', 'Hydraulic Oil 46 (20L)', 'CAN', 2800.00, 15.0, 'GL-5003', 'CC-MAINT'),
    ('MAT-013', 'Cable Tie 300mm (100 pcs)', 'PKT', 120.00, 50.0, 'GL-5002', 'CC-ELEC'),
    ('MAT-014', 'Welding Rod 3.15mm (5kg)', 'PKT', 550.00, 40.0, 'GL-5004', 'CC-WELDING'),
    ('MAT-015', 'Safety Helmet (Blue)', 'NOS', 350.00, 30.0, 'GL-5005', 'CC-SAFETY'),
    ('MAT-016', 'Cotton Waste', 'KG', 35.00, 500.0, 'GL-5003', 'CC-MAINT'),
    ('MAT-017', 'Motor Oil SAE 30 (5L)', 'CAN', 650.00, 20.0, 'GL-5003', 'CC-MAINT'),
    ('MAT-018', 'Hacksaw Blade 12 inch', 'NOS', 25.00, 200.0, 'GL-5001', 'CC-MECH'),
    ('MAT-019', 'MCB 32A Single Pole', 'NOS', 285.00, 40.0, 'GL-5002', 'CC-ELEC'),
    ('MAT-020', 'Teflon Tape (12mm x 12m)', 'NOS', 18.00, 500.0, 'GL-5001', 'CC-MECH'),
]

for row_num, row_data in enumerate(data, 2):
    for col_num, value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=value)

# Auto-size columns
for col in ws.columns:
    max_length = max(len(str(cell.value or '')) for cell in col)
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 4, 50)

wb.save('sample_inventory.xlsx')
print("Sample inventory file created: sample_inventory.xlsx")
print("Upload this file using the Store Manager account to populate inventory.")
